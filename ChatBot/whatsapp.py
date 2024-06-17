from twilio.rest import Client  # pip3 install twilio
import openpyxl

account_sid = ''
auth_token = ''
client = Client(account_sid, auth_token)


def sendMessage():
    workbook = openpyxl.load_workbook('Testdata.xlsx')
    sheet = workbook['whatsappnumbers']
    max_column = sheet.max_column
    max_row = sheet.max_row
    for row in range(2, max_row+1):
        name = sheet.cell(row=row, column=1)
        to_number = sheet.cell(row=row, column=2)
        message = sheet.cell(row=row, column=3)
        message = client.messages.create(
            body=message.value,
            from_='whatsapp:+14155238886',  # This is the Twilio sandbox number
            to='whatsapp:'+to_number.value
        )
    print('Reminder message sent to ' + name.value)
